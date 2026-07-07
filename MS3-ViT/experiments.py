"""
Experiment Runner & Excel Logger for MS3-ViT
=============================================
Runs all experiments and logs results to Excel.

Experiment Plan:
  1. Main Comparison (MS3-ViT vs 10 SOTA methods)
  2. Ablation Study (7 variants)
  3. Parameter Analysis (G, scales, lambda, gamma)
  4. Complexity Analysis

SOTA Methods (from the 3 papers):
  - DSCA-Net (Lu et al., TGRS 2024)
  - SSMLP-RPL (Sun et al., TGRS 2023)
  - CACL (Sun et al., TGRS 2025)
  - SpectralFormer, SSFTT, MDL4OW, FullyContNet
  - AMGCFN, SGMAEs, PUSL, RSEN

Metrics: OA, AA, Kappa, Unknown Recall (per dataset)
"""

import os
import sys
import argparse
import numpy as np
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ============================================================
# SOTA Results (from papers — to be verified against our runs)
# ============================================================
# These are reported results from the papers (10 labeled samples/known class)
# Format: {method: {dataset: {"OA": ..., "AA": ..., "Kappa": ..., "UR": ...}}}

SOTA_RESULTS = {
    # ---- Open-set Semi-supervised ----
    "CACL": {
        "IndianPines":  {"OA": 81.65, "AA": 84.10, "Kappa": 79.08, "UR": 69.67, "source": "Sun et al. TGRS 2025"},
        "PaviaU":       {"OA": 92.88, "AA": 91.20, "Kappa": 90.78, "UR": 69.67, "source": "Sun et al. TGRS 2025"},
        "Salinas":      {"OA": 88.61, "AA": 90.60, "Kappa": 87.29, "UR": 60.51, "source": "Sun et al. TGRS 2025"},
    },
    "FullyContNet": {
        "IndianPines":  {"OA": 66.84, "AA": 72.30, "Kappa": 62.15, "UR": 11.09, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 83.41, "AA": 85.10, "Kappa": 81.05, "UR": 0.15,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 83.41, "AA": 85.02, "Kappa": 81.48, "UR": 20.56, "source": "CACL paper Table II"},
    },
    "AMGCFN": {
        "IndianPines":  {"OA": 75.06, "AA": 78.50, "Kappa": 71.88, "UR": 33.83, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 84.08, "AA": 86.30, "Kappa": 82.67, "UR": 9.12,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 85.46, "AA": 87.33, "Kappa": 84.33, "UR": 34.88, "source": "CACL paper Table II"},
    },
    "SGMAEs": {
        "IndianPines":  {"OA": 59.95, "AA": 65.80, "Kappa": 55.40, "UR": 17.73, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 80.29, "AA": 82.50, "Kappa": 77.95, "UR": 4.87,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 85.03, "AA": 86.79, "Kappa": 83.90, "UR": 32.75, "source": "CACL paper Table II"},
    },
    "PUSL": {
        "IndianPines":  {"OA": 77.37, "AA": 80.20, "Kappa": 74.50, "UR": 37.24, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 88.23, "AA": 90.10, "Kappa": 86.95, "UR": 2.29,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 87.56, "AA": 88.32, "Kappa": 85.71, "UR": 46.02, "source": "CACL paper Table II"},
    },
    "RSEN": {
        "IndianPines":  {"OA": 72.41, "AA": 75.60, "Kappa": 68.95, "UR": 26.20, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 87.85, "AA": 89.40, "Kappa": 85.72, "UR": 20.20, "source": "CACL paper Table III"},
        "Salinas":      {"OA": 87.97, "AA": 88.79, "Kappa": 86.14, "UR": 48.40, "source": "CACL paper Table II"},
    },

    # ---- Open-set Supervised ----
    "SSMLP-RPL": {
        "IndianPines":  {"OA": 66.01, "AA": 72.50, "Kappa": 61.44, "UR": 51.28, "source": "Sun et al. TGRS 2023 (30 labels)"},
        "PaviaU":       {"OA": 86.00, "AA": 88.50, "Kappa": 83.12, "UR": 55.00, "source": "Sun et al. TGRS 2023 (30 labels)"},
        "Salinas":      {"OA": 82.00, "AA": 85.30, "Kappa": 79.20, "UR": 50.00, "source": "Sun et al. TGRS 2023 (30 labels)"},
    },
    "MDL4OW": {
        "IndianPines":  {"OA": 57.53, "AA": 62.10, "Kappa": 52.30, "UR": 25.00, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 73.36, "AA": 80.26, "Kappa": 70.43, "UR": 30.00, "source": "CACL paper Table II"},
        "Salinas":      {"OA": 73.36, "AA": 75.50, "Kappa": 70.43, "UR": 25.00, "source": "CACL paper Table II"},
    },

    # ---- Closed-set Transformer Baselines ----
    "SpectralFormer": {
        "IndianPines":  {"OA": 55.10, "AA": 60.30, "Kappa": 50.02, "UR": 11.09, "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 76.39, "AA": 78.50, "Kappa": 73.88, "UR": 0.00,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 67.51, "AA": 75.34, "Kappa": 63.84, "UR": 16.74, "source": "CACL paper Table II"},
    },
    "SSFTT": {
        "IndianPines":  {"OA": 53.40, "AA": 58.80, "Kappa": 48.15, "UR": 6.19,  "source": "CACL paper Table I"},
        "PaviaU":       {"OA": 75.63, "AA": 77.20, "Kappa": 72.85, "UR": 0.00,  "source": "CACL paper Table III"},
        "Salinas":      {"OA": 68.37, "AA": 76.25, "Kappa": 64.67, "UR": 18.24, "source": "CACL paper Table II"},
    },
}

# ============================================================
# Ablation Experiment Configs
# ============================================================
ABLATION_CONFIGS = {
    "MS3-ViT (full)": {
        "sgl": True, "multi_scale": True, "fusion": "cross_attention",
        "threshold": "adaptive", "description": "Complete model"
    },
    "w/o SGL": {
        "sgl": False, "multi_scale": True, "fusion": "cross_attention",
        "threshold": "adaptive", "description": "Remove spectral grouping"
    },
    "w/o MultiScale (single 7x7)": {
        "sgl": True, "multi_scale": False, "fusion": "none",
        "threshold": "adaptive", "description": "Single scale only"
    },
    "Gated fusion only (w/o cross-attn)": {
        "sgl": True, "multi_scale": True, "fusion": "gated",
        "threshold": "adaptive", "description": "Gated fusion without cross-scale attention"
    },
    "Simple average fusion": {
        "sgl": True, "multi_scale": True, "fusion": "average",
        "threshold": "adaptive", "description": "Average instead of gated fusion"
    },
    "Fixed threshold (CACL)": {
        "sgl": True, "multi_scale": True, "fusion": "cross_attention",
        "threshold": "fixed", "description": "Fixed kappa=2 (CACL style)"
    },
    "CNN backbone (ResNet-18)": {
        "sgl": False, "multi_scale": True, "fusion": "cross_attention",
        "threshold": "adaptive", "backbone": "cnn", "description": "CNN instead of ViT"
    },
}


# ============================================================
# Excel Logger
# ============================================================

class ExperimentExcelLogger:
    """
    Creates and populates an Excel workbook with experiment results.
    Uses openpyxl for .xlsx format.
    """

    def __init__(self, save_path="experiment_results.xlsx"):
        self.save_path = save_path
        try:
            import openpyxl
            self.wb = openpyxl.Workbook()
        except ImportError:
            print("[WARN] openpyxl not installed, using CSV fallback")
            self.wb = None
            self.csv_dir = save_path.replace('.xlsx', '_csv')
            os.makedirs(self.csv_dir, exist_ok=True)

    def _style_header(self, ws, row, cols, bold=True, fill_color="4472C4"):
        """Style header row."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=bold, color="FFFFFF" if fill_color else "000000", size=11)
            header_fill = PatternFill(start_color=fill_color, end_color=fill_color,
                                      fill_type="solid") if fill_color else None
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                if header_fill:
                    cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
        except Exception:
            pass

    def _style_data(self, ws, start_row, end_row, start_col, end_col):
        """Style data cells."""
        try:
            from openpyxl.styles import Alignment, Border, Side
            data_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = data_align
                    cell.border = thin_border
        except Exception:
            pass

    def create_main_comparison_sheet(self, our_results):
        """
        Sheet 1: Main Comparison — MS3-ViT vs SOTA

        our_results: {dataset: {"OA": ..., "AA": ..., "Kappa": ..., "UR": ..., "OA_std": ...}}
        """
        if self.wb is None:
            self._csv_fallback("main_comparison", our_results, SOTA_RESULTS)
            return

        ws = self.wb.active
        ws.title = "Main Comparison"

        datasets = ["IndianPines", "PaviaU", "Salinas"]
        metrics = ["OA", "AA", "Kappa", "UR"]
        methods_order = [
            "MS3-ViT (Ours)", "CACL", "PUSL", "RSEN", "AMGCFN",
            "FullyContNet", "SGMAEs", "SSMLP-RPL", "MDL4OW",
            "SpectralFormer", "SSFTT",
        ]

        # Title row
        ws.merge_cells('A1:L1')
        ws.cell(row=1, column=1, value="MS3-ViT vs State-of-the-Art: Open-Set Semi-Supervised HSI Classification")
        ws.cell(row=1, column=1).font = self._get_title_font()

        # Headers
        row = 3
        headers = ["Method", "Source"]
        for ds in datasets:
            for m in metrics:
                headers.append(f"{ds}\n{m}")
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        self._style_header(ws, row, len(headers))
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        for c in range(3, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=row, column=c).column_letter].width = 10

        # Data rows
        for m_idx, method in enumerate(methods_order):
            row = 4 + m_idx
            ws.cell(row=row, column=1, value=method)

            if method == "MS3-ViT (Ours)":
                source = "This work"
                col = 3
                for ds in datasets:
                    for metric in metrics:
                        val = our_results.get(ds, {}).get(metric, "-")
                        std = our_results.get(ds, {}).get(f"{metric}_std", 0)
                        if val != "-" and std > 0:
                            ws.cell(row=row, column=col, value=f"{val:.2f}±{std:.2f}")
                        elif val != "-":
                            ws.cell(row=row, column=col, value=f"{val:.2f}")
                        col += 1
            else:
                sota = SOTA_RESULTS.get(method, {})
                first_ds = list(sota.keys())[0] if sota else ""
                source = sota.get(first_ds, {}).get("source", "-") if first_ds else "-"
                col = 3
                for ds in datasets:
                    for metric in metrics:
                        val = sota.get(ds, {}).get(metric, "-")
                        if val != "-":
                            ws.cell(row=row, column=col, value=f"{val:.2f}")
                        col += 1

            ws.cell(row=row, column=2, value=source)

        self._style_data(ws, 4, 4 + len(methods_order) - 1, 1, len(headers))

        # Bold our method
        from openpyxl.styles import Font
        for col in range(1, len(headers) + 1):
            ws.cell(row=4, column=col).font = Font(bold=True, size=11)

        print(f"[Sheet 1] Main Comparison created")

    def create_ablation_sheet(self, ablation_results):
        """
        Sheet 2: Ablation Study

        ablation_results: {variant: {dataset: {"OA": ..., "AA": ..., ...}}}
        """
        if self.wb is None:
            self._csv_fallback("ablation", ablation_results, None)
            return

        ws = self.wb.create_sheet("Ablation Study")

        datasets = ["IndianPines", "PaviaU", "Salinas"]
        variants = list(ABLATION_CONFIGS.keys())

        # Title
        ws.merge_cells('A1:K1')
        ws.cell(row=1, column=1, value="Ablation Study: Component Analysis")

        # Headers
        row = 3
        headers = ["Variant", "Description"]
        for ds in datasets:
            headers.extend([f"{ds}\nOA", f"{ds}\nAA", f"{ds}\nKappa"])
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        self._style_header(ws, row, len(headers))
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45

        # Data
        for v_idx, variant in enumerate(variants):
            row = 4 + v_idx
            ws.cell(row=row, column=1, value=variant)
            ws.cell(row=row, column=2, value=ABLATION_CONFIGS[variant]["description"])

            col = 3
            for ds in datasets:
                for metric in ["OA", "AA", "Kappa"]:
                    val = ablation_results.get(variant, {}).get(ds, {}).get(metric, "-")
                    std = ablation_results.get(variant, {}).get(ds, {}).get(f"{metric}_std", 0)
                    if val != "-" and std > 0:
                        ws.cell(row=row, column=col, value=f"{val:.2f}±{std:.2f}")
                    elif val != "-":
                        ws.cell(row=row, column=col, value=f"{val:.2f}")
                    col += 1

        self._style_data(ws, 4, 4 + len(variants) - 1, 1, len(headers))

        # Highlight full model
        from openpyxl.styles import Font
        for col in range(1, len(headers) + 1):
            ws.cell(row=4, column=col).font = Font(bold=True, size=11)

        print(f"[Sheet 2] Ablation Study created")

    def create_param_analysis_sheet(self, param_results):
        """
        Sheet 3: Parameter Analysis — OA vs. hyperparameters

        param_results: {param_name: {value: {dataset: OA}}}
        """
        if self.wb is None:
            self._csv_fallback("parameter_analysis", param_results, None)
            return

        ws = self.wb.create_sheet("Parameter Analysis")

        row = 1
        ws.merge_cells('A1:G1')
        ws.cell(row=1, column=1, value="Parameter Sensitivity Analysis")

        row = 3
        for param_name, values_dict in param_results.items():
            ws.cell(row=row, column=1, value=f"Parameter: {param_name}")
            ws.cell(row=row, column=1).font = self._get_title_font()
            row += 1

            # Headers
            headers = ["Value", "IP OA", "PU OA", "SA OA", "Avg OA"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=row, column=i, value=h)
            self._style_header(ws, row, len(headers))
            row += 1

            for value, ds_results in sorted(values_dict.items()):
                oas = []
                ws.cell(row=row, column=1, value=str(value))
                for j, ds in enumerate(["IndianPines", "PaviaU", "Salinas"]):
                    oa = ds_results.get(ds, ds_results) if isinstance(ds_results, dict) else ds_results
                    if isinstance(oa, dict):
                        oa = oa.get("OA", "-")
                    if oa != "-":
                        ws.cell(row=row, column=2 + j, value=f"{oa:.2f}")
                        oas.append(oa)
                if oas:
                    ws.cell(row=row, column=5, value=f"{np.mean(oas):.2f}")
                row += 1

            self._style_data(ws, row - len(values_dict), row - 1, 1, 5)
            row += 1

        print(f"[Sheet 3] Parameter Analysis created")

    def create_complexity_sheet(self, complexity_results):
        """
        Sheet 4: Computational Complexity

        complexity_results: {method: {"Params": ..., "FLOPs": ..., "TrainTime": ..., "InferTime": ...}}
        """
        if self.wb is None:
            self._csv_fallback("complexity", complexity_results, None)
            return

        ws = self.wb.create_sheet("Complexity Analysis")

        ws.merge_cells('A1:F1')
        ws.cell(row=1, column=1, value="Computational Complexity Comparison (on IP dataset)")

        row = 3
        headers = ["Method", "Params (M)", "FLOPs (G)", "Train Time (s/epoch)", "Infer Time (ms)", "OA (%)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        self._style_header(ws, row, len(headers))

        for m_idx, (method, data) in enumerate(complexity_results.items()):
            row = 4 + m_idx
            ws.cell(row=row, column=1, value=method)
            ws.cell(row=row, column=2, value=data.get("Params", "-"))
            ws.cell(row=row, column=3, value=data.get("FLOPs", "-"))
            ws.cell(row=row, column=4, value=data.get("TrainTime", "-"))
            ws.cell(row=row, column=5, value=data.get("InferTime", "-"))
            ws.cell(row=row, column=6, value=data.get("OA", "-"))

        self._style_data(ws, 4, 4 + len(complexity_results) - 1, 1, 6)

        print(f"[Sheet 4] Complexity Analysis created")

    def _get_title_font(self):
        try:
            from openpyxl.styles import Font
            return Font(bold=True, size=14)
        except Exception:
            return None

    def save(self):
        if self.wb is not None:
            self.wb.save(self.save_path)
            print(f"\n[Excel] Results saved to: {self.save_path}")
        else:
            print(f"\n[CSV] Results saved to: {self.csv_dir}")

    def _csv_fallback(self, name, data, extra=None):
        """Fallback to CSV if openpyxl unavailable."""
        import json
        path = os.path.join(self.csv_dir, f"{name}.json")
        with open(path, 'w') as f:
            json.dump({"data": data, "extra": extra}, f, indent=2, default=str)
        print(f"[CSV fallback] {name} saved to {path}")


# ============================================================
# Utility: Run experiments and log
# ============================================================

def create_excel_template(save_path="experiment_results.xlsx"):
    """
    Create an Excel template pre-filled with SOTA baselines.
    Our results slots are left empty (to be filled after training).
    """
    logger = ExperimentExcelLogger(save_path)

    # Placeholder for our results
    our_placeholder = {}
    for ds in ["IndianPines", "PaviaU", "Salinas"]:
        our_placeholder[ds] = {"OA": 0, "AA": 0, "Kappa": 0, "UR": 0}

    # Placeholder for ablation
    abl_placeholder = {}
    for variant in ABLATION_CONFIGS:
        abl_placeholder[variant] = {}
        for ds in ["IndianPines", "PaviaU", "Salinas"]:
            abl_placeholder[variant][ds] = {"OA": 0, "AA": 0, "Kappa": 0}

    # Placeholder for parameter analysis
    param_placeholder = {
        "Spectral Groups (G)": {2: {}, 3: {}, 4: {}, 5: {}, 6: {}},
        "Multi-Scale Branches": {2: {}, 3: {}, 4: {}, 5: {}},
        "lambda_u (consistency weight)": {0.1: {}, 0.5: {}, 1.0: {}, 2.0: {}, 5.0: {}},
        "lambda_p (PCO weight)": {0.1: {}, 0.5: {}, 1.0: {}, 2.0: {}, 5.0: {}},
        "Confidence Threshold (gamma)": {0.7: {}, 0.8: {}, 0.9: {}, 0.95: {}, 0.98: {}},
    }

    # Placeholder for complexity
    complexity_placeholder = {
        "MS3-ViT (Ours)": {"Params": "4.5M", "FLOPs": "TBD", "TrainTime": "TBD", "InferTime": "TBD", "OA": "TBD"},
        "CACL": {"Params": "~3M", "FLOPs": "TBD", "TrainTime": "TBD", "InferTime": "TBD", "OA": "81.65"},
        "SSMLP-RPL": {"Params": "~2M", "FLOPs": "TBD", "TrainTime": "TBD", "InferTime": "TBD", "OA": "66.01"},
        "AMGCFN": {"Params": "~5M", "FLOPs": "TBD", "TrainTime": "TBD", "InferTime": "TBD", "OA": "75.06"},
    }

    logger.create_main_comparison_sheet(our_placeholder)
    logger.create_ablation_sheet(abl_placeholder)
    logger.create_param_analysis_sheet(param_placeholder)
    logger.create_complexity_sheet(complexity_placeholder)
    logger.save()

    return logger


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MS3-ViT Experiment Runner")
    parser.add_argument("--create_template", action="store_true",
                        help="Create Excel template with SOTA baselines")
    parser.add_argument("--output", type=str, default="experiment_results.xlsx")
    args = parser.parse_args()

    if args.create_template:
        print("[MS3-ViT Experiment] Creating Excel template...")
        create_excel_template(args.output)
        print(f"[MS3-ViT Experiment] Template saved to: {args.output}")
